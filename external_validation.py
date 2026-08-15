"""
external_validation.py

외부 문헌 검증셋(polymer_external_literature_validation_master.xlsx)에
학습된 5-fold 앙상블 모델로 예측을 채워넣고, 문헌값과 비교하는 스크립트.

STEP12(Grad-CAM) 이후에 붙는 "외부 검증" 스텝.
재학습/파이프라인 변경 없음 -- 기존 confirmation_checkpoints만 사용.

실행: python external_validation.py
(프로젝트 루트에서, app.py와 같은 레벨. app.py의 모델 로딩 로직을 그대로 재사용)

입력: polymer_external_literature_validation_master.xlsx
출력: polymer_external_literature_validation_FILLED.xlsx
      (같은 폴더, 원본은 건드리지 않음)
"""

import sys
from pathlib import Path

import numpy as np
import openpyxl
import torch
from PIL import Image

sys.path.insert(0, str(Path(__file__).parent / "ensemble"))
from ensemble_common import build_final_config, load_fold_model, load_pool_and_metadata, get_fold_splits  # noqa: E402
from datasets.dataset import LABEL_ORDER, compute_fold_stats  # noqa: E402
from datasets.transforms import get_valid_transform  # noqa: E402

from preprocessing.smiles_to_mol import parse_smiles  # noqa: E402
from preprocessing.mol_to_image import render_mol_to_png, IMAGE_SIZE  # noqa: E402

# ---------------------------------------------------------------------------
# 설정
# ---------------------------------------------------------------------------
N_FOLDS = 5
INPUT_PATH = Path("polymer_external_literature_validation_master.xlsx")
OUTPUT_PATH = Path("polymer_external_literature_validation_FILLED.xlsx")

# 검증 대상 시트 (SMILES/문헌값 있는 시트만)
TARGET_SHEETS = ["HighQuality_Panel", "Master_Literature"]

# 시트별 컬럼명 -> LABEL_ORDER 인덱스 매핑용 (property 이름 그대로 일치)
# AI_Tg, AI_Density, AI_FFV, AI_Tc, AI_Rg 컬럼이 있다고 가정
AI_COL_PREFIX = "AI_"


# ---------------------------------------------------------------------------
# 1. 모델 로드 (app.py와 동일 로직)
# ---------------------------------------------------------------------------
def load_models():
    print("모델 로딩 중 (5-fold)...")
    config = build_final_config()
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

    pool_df, _ = load_pool_and_metadata(config)
    splits = dict(get_fold_splits(pool_df))

    models = {}
    label_stats = {}
    for fold in range(1, N_FOLDS + 1):
        train_idx, _ = splits[fold]
        train_df = pool_df.iloc[train_idx].reset_index(drop=True)
        label_mean, label_std = compute_fold_stats(train_df, LABEL_ORDER)
        label_stats[fold] = (label_mean, label_std)
        models[fold] = load_fold_model(fold, config, device)
        print(f"  fold{fold} 로드 완료")

    image_cfg = config["image"]
    transform = get_valid_transform(
        image_cfg["size"], image_cfg["normalize_mean"], image_cfg["normalize_std"]
    )
    print(f"모델 로딩 완료. device={device}")
    return models, label_stats, transform, device


# ---------------------------------------------------------------------------
# 2. SMILES -> 예측 (5-fold 앙상블)
# ---------------------------------------------------------------------------
def predict_smiles(smiles, models, label_stats, transform, device):
    """반환: dict{property_name: value} 또는 None (파싱/렌더링 실패 시)"""
    mol = parse_smiles(smiles)
    if mol is None:
        return None, "parse_fail"

    import tempfile
    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp_path = Path(tmp_dir) / "render.png"
        ok = render_mol_to_png(mol, tmp_path, size=IMAGE_SIZE)
        if not ok:
            return None, "render_fail"
        img = Image.open(tmp_path).convert("RGB")
        img.load()

    image_np = np.array(img)
    image_tensor = transform(image=image_np)["image"].unsqueeze(0).to(device)

    fold_preds = []
    for fold in range(1, N_FOLDS + 1):
        model = models[fold]
        label_mean, label_std = label_stats[fold]
        mean_arr = np.array([label_mean[name] for name in LABEL_ORDER])
        std_arr = np.array([label_std[name] for name in LABEL_ORDER])
        with torch.no_grad():
            out = model(image_tensor).cpu().numpy()[0]
        fold_preds.append(out * std_arr + mean_arr)

    ensemble = np.mean(fold_preds, axis=0)
    return {name: float(val) for name, val in zip(LABEL_ORDER, ensemble)}, None


# ---------------------------------------------------------------------------
# 3. 시트 처리: 헤더 찾기 -> 각 row의 Repeat_SMILES 예측 -> AI_*/AbsErr_*/In_Range 채우기
# ---------------------------------------------------------------------------
def process_sheet(ws, models, label_stats, transform, device):
    # 헤더 행 찾기 (Repeat_SMILES가 있는 행)
    header_row_idx = None
    headers = None
    for r in range(1, 6):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if "Repeat_SMILES" in row_vals:
            header_row_idx = r
            headers = row_vals
            break
    if header_row_idx is None:
        print(f"  [스킵] '{ws.title}': Repeat_SMILES 컬럼 못 찾음")
        return 0, 0

    col_idx = {name: i + 1 for i, name in enumerate(headers) if name}

    smiles_col = col_idx.get("Repeat_SMILES")
    ok_count, fail_count = 0, 0

    for r in range(header_row_idx + 1, ws.max_row + 1):
        smiles = ws.cell(row=r, column=smiles_col).value
        if not smiles or not str(smiles).strip():
            continue
        smiles = str(smiles).strip()

        preds, err = predict_smiles(smiles, models, label_stats, transform, device)
        if preds is None:
            fail_count += 1
            print(f"  [실패] row {r}: {smiles[:40]}... ({err})")
            continue
        ok_count += 1

        for prop in LABEL_ORDER:
            ai_col_name = f"{AI_COL_PREFIX}{prop}"
            if ai_col_name in col_idx:
                ws.cell(row=r, column=col_idx[ai_col_name], value=round(preds[prop], 4))

            # AbsErr_{prop}: Lit_{prop} 또는 Lit_{prop}_C 컬럼과 비교 (있는 경우만)
            for lit_col_candidate in (f"Lit_{prop}", f"Lit_{prop}_C"):
                if lit_col_candidate in col_idx:
                    lit_val = ws.cell(row=r, column=col_idx[lit_col_candidate]).value
                    abserr_col = f"AbsErr_{prop}"
                    if lit_val is not None and abserr_col in col_idx:
                        try:
                            abserr = abs(float(lit_val) - preds[prop])
                            ws.cell(row=r, column=col_idx[abserr_col], value=round(abserr, 4))
                        except (TypeError, ValueError):
                            pass
                    break

            # {prop}_In_Range: Lit_{prop}_Min / Lit_{prop}_Max 범위 안에 드는지
            in_range_col = f"{prop}_In_Range"
            min_col = f"{prop}_Min" if f"{prop}_Min" in col_idx else f"Lit_{prop}_Min"
            max_col = f"{prop}_Max" if f"{prop}_Max" in col_idx else f"Lit_{prop}_Max"
            if in_range_col in col_idx and min_col in col_idx and max_col in col_idx:
                vmin = ws.cell(row=r, column=col_idx[min_col]).value
                vmax = ws.cell(row=r, column=col_idx[max_col]).value
                if vmin is not None and vmax is not None:
                    try:
                        in_range = float(vmin) <= preds[prop] <= float(vmax)
                        ws.cell(row=r, column=col_idx[in_range_col], value=in_range)
                    except (TypeError, ValueError):
                        pass

    return ok_count, fail_count


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------
def main():
    models, label_stats, transform, device = load_models()

    wb = openpyxl.load_workbook(INPUT_PATH)

    total_ok, total_fail = 0, 0
    for sheet_name in TARGET_SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"[스킵] 시트 없음: {sheet_name}")
            continue
        print(f"\n=== {sheet_name} 처리 중 ===")
        ok, fail = process_sheet(wb[sheet_name], models, label_stats, transform, device)
        total_ok += ok
        total_fail += fail
        print(f"  완료: 성공 {ok} / 실패 {fail}")

    wb.save(OUTPUT_PATH)
    print(f"\n저장 완료: {OUTPUT_PATH}")
    print(f"전체: 성공 {total_ok} / 실패 {total_fail}")


if __name__ == "__main__":
    main()
